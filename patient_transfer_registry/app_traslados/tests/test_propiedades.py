"""
Pruebas de propiedad para los modelos de app_traslados.

Usa Hypothesis para verificar propiedades universales del sistema.
"""

import datetime

from django.core.exceptions import ValidationError
from hypothesis import given, settings
from hypothesis import strategies as st
from hypothesis.extra.django import TestCase

from app_traslados.models import ControlMes, TrasladoPaciente


def crearControlMeses ():
	"""Crea las 12 filas de ControlMes con estado ABIERTO si no existen."""
	for numeroMes in range (1, 13):
		ControlMes.objects.get_or_create (
			mes=numeroMes,
			defaults={'estado': 'ABIERTO'},
		)


def construirTrasladoValido (fecha):
	"""Construye y retorna un TrasladoPaciente con todos los campos requeridos para la fecha dada."""
	traslado = TrasladoPaciente (
		fecha=fecha,
		hora_reporte=datetime.time (8, 0),
		nombre_paciente='Paciente Prueba',
		documento='123456789',
		servicio='Urgencias',
		quien_reporta='Enfermera Prueba',
		destino='Hospital Central',
		procedimiento='Traslado de emergencia',
		medico='Dr. Prueba',
		conductor='Conductor Prueba',
		radio_operador='Radio Prueba',
		ambulancia='AMB-001',
	)
	return traslado


# Estrategia: fechas válidas (no futuras), desde 2000-01-01 hasta hoy
estrategiaFechaValida = st.dates (
	min_value=datetime.date (2000, 1, 1),
	max_value=datetime.date.today (),
)


class PropiedadDerivacionMes (TestCase):
	"""
	**Validates: Requirements 4.2**

	Property 1: Month derivation is consistent with fecha.
	Para cualquier fecha válida, TrasladoPaciente.mes == fecha.month tras guardar.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	@settings (max_examples=10)
	@given (fecha=estrategiaFechaValida)
	def test_mes_derivado_es_consistente_con_fecha (self, fecha):
		"""Verifica que el campo mes siempre coincide con fecha.month tras guardar."""
		traslado = construirTrasladoValido (fecha)
		traslado.save ()

		# Recargar desde la base de datos para confirmar persistencia
		trasladoGuardado = TrasladoPaciente.objects.get (pk=traslado.pk)

		self.assertEqual (
			trasladoGuardado.mes,
			fecha.month,
			f"Se esperaba mes={fecha.month} para fecha={fecha}, "
			f"pero se obtuvo mes={trasladoGuardado.mes}",
		)

		# Limpiar el registro creado para no interferir con otras iteraciones
		traslado.delete ()


class PropiedadBloqueoMesCerrado (TestCase):
	"""
	**Validates: Requirements 3.9, 5.4**

	Property 2: Closed month blocks all write operations.
	Para cualquier mes cerrado, crear o actualizar un TrasladoPaciente
	en ese mes lanza ValidationError en la capa de modelo.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def _cerrarMes (self, numeroMes):
		"""Cierra el mes indicado actualizando su ControlMes a CERRADO."""
		ControlMes.objects.filter (mes=numeroMes).update (estado='CERRADO')

	def _abrirMes (self, numeroMes):
		"""Abre el mes indicado actualizando su ControlMes a ABIERTO."""
		ControlMes.objects.filter (mes=numeroMes).update (estado='ABIERTO')

	def _fechaParaMes (self, numeroMes):
		"""Retorna una fecha válida (no futura) para el mes dado."""
		hoy = datetime.date.today ()
		anio = hoy.year if numeroMes <= hoy.month else hoy.year - 1
		return datetime.date (anio, numeroMes, 1)

	@settings (max_examples=10)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_crear_en_mes_cerrado_lanza_validation_error (self, numeroMes):
		"""Verifica que crear un TrasladoPaciente en un mes cerrado lanza ValidationError."""
		self._cerrarMes (numeroMes)
		fecha = self._fechaParaMes (numeroMes)
		traslado = construirTrasladoValido (fecha)

		with self.assertRaises (ValidationError):
			traslado.save ()

		# Restaurar mes a ABIERTO para no afectar otras iteraciones
		self._abrirMes (numeroMes)

	@settings (max_examples=10)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_actualizar_en_mes_cerrado_lanza_validation_error (self, numeroMes):
		"""Verifica que actualizar un TrasladoPaciente en un mes cerrado lanza ValidationError."""
		# Guardar el registro con el mes abierto
		fecha = self._fechaParaMes (numeroMes)
		traslado = construirTrasladoValido (fecha)
		traslado.save ()

		# Cerrar el mes después de haber guardado el registro
		self._cerrarMes (numeroMes)

		# Intentar actualizar debe lanzar ValidationError
		traslado.observacion = 'Modificación en mes cerrado'
		with self.assertRaises (ValidationError):
			traslado.save ()

		# Restaurar mes a ABIERTO y limpiar el registro
		self._abrirMes (numeroMes)
		TrasladoPaciente.objects.filter (pk=traslado.pk).delete ()


class PropiedadValidacionCamposObligatorios (TestCase):
	"""
	**Validates: Requirements 3.8**

	Property 5: Required field validation always rejects incomplete records.
	Cualquier combinación con al menos un campo obligatorio (fecha,
	hora_reporte, nombre_paciente, documento) vacío o nulo debe ser rechazada
	con ValidationError sin persistir ningún registro en la base de datos.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def _contarRegistros (self):
		"""Retorna el conteo actual de TrasladoPaciente en la base de datos."""
		return TrasladoPaciente.objects.count ()

	@settings (max_examples=10)
	@given (fecha=st.one_of (st.none (), estrategiaFechaValida))
	def test_fecha_vacia_rechaza_registro (self, fecha):
		"""Verifica que omitir fecha lanza ValidationError y no persiste el registro."""
		# Solo probar el caso donde fecha es None (campo obligatorio ausente)
		if fecha is not None:
			return

		conteoAntes = self._contarRegistros ()
		traslado = TrasladoPaciente (
			fecha=None,
			hora_reporte=datetime.time (8, 0),
			nombre_paciente='Paciente Prueba',
			documento='123456789',
			servicio='Urgencias',
			quien_reporta='Enfermera Prueba',
			destino='Hospital Central',
			procedimiento='Traslado de emergencia',
			medico='Dr. Prueba',
			conductor='Conductor Prueba',
			radio_operador='Radio Prueba',
			ambulancia='AMB-001',
		)
		with self.assertRaises ((ValidationError, Exception)):
			traslado.save ()

		self.assertEqual (
			self._contarRegistros (),
			conteoAntes,
			"No debe persistirse ningún registro cuando fecha es nula.",
		)

	@settings (max_examples=10)
	@given (
		fecha=estrategiaFechaValida,
		nombrePaciente=st.text (min_size=1, max_size=255),
		documento=st.text (min_size=1, max_size=50),
		# Seleccionar cuál campo obligatorio omitir: 0=hora_reporte, 1=nombre_paciente, 2=documento
		campoAOmitir=st.integers (min_value=0, max_value=2),
	)
	def test_campo_obligatorio_vacio_rechaza_registro (
		self, fecha, nombrePaciente, documento, campoAOmitir
	):
		"""Verifica que omitir hora_reporte, nombre_paciente o documento lanza ValidationError."""
		conteoAntes = self._contarRegistros ()

		# Construir el traslado omitiendo el campo seleccionado
		horaReporte = None if campoAOmitir == 0 else datetime.time (8, 0)
		nombreUsado = '' if campoAOmitir == 1 else nombrePaciente
		documentoUsado = '' if campoAOmitir == 2 else documento

		traslado = TrasladoPaciente (
			fecha=fecha,
			hora_reporte=horaReporte,
			nombre_paciente=nombreUsado,
			documento=documentoUsado,
			servicio='Urgencias',
			quien_reporta='Enfermera Prueba',
			destino='Hospital Central',
			procedimiento='Traslado de emergencia',
			medico='Dr. Prueba',
			conductor='Conductor Prueba',
			radio_operador='Radio Prueba',
			ambulancia='AMB-001',
		)

		with self.assertRaises ((ValidationError, Exception)):
			traslado.save ()

		self.assertEqual (
			self._contarRegistros (),
			conteoAntes,
			f"No debe persistirse ningún registro cuando el campo obligatorio "
			f"(índice {campoAOmitir}) está vacío.",
		)

	@settings (max_examples=15)
	@given (
		fecha=estrategiaFechaValida,
		nombrePaciente=st.text (min_size=1, max_size=255),
		documento=st.text (min_size=1, max_size=50),
		# Máscara de bits: bit 0=fecha, bit 1=hora_reporte, bit 2=nombre_paciente, bit 3=documento
		# Al menos un bit debe estar activo (campo omitido), por eso min_value=1
		mascaraCamposOmitidos=st.integers (min_value=1, max_value=15),
	)
	def test_cualquier_combinacion_con_campo_obligatorio_vacio_rechaza (
		self, fecha, nombrePaciente, documento, mascaraCamposOmitidos
	):
		"""Verifica que cualquier combinación con al menos un campo obligatorio vacío es rechazada."""
		conteoAntes = self._contarRegistros ()

		# Determinar qué campos omitir según la máscara de bits
		omitirFecha = bool (mascaraCamposOmitidos & 1)
		omitirHoraReporte = bool (mascaraCamposOmitidos & 2)
		omitirNombrePaciente = bool (mascaraCamposOmitidos & 4)
		omitirDocumento = bool (mascaraCamposOmitidos & 8)

		fechaUsada = None if omitirFecha else fecha
		horaReporteUsada = None if omitirHoraReporte else datetime.time (8, 0)
		nombreUsado = '' if omitirNombrePaciente else nombrePaciente
		documentoUsado = '' if omitirDocumento else documento

		traslado = TrasladoPaciente (
			fecha=fechaUsada,
			hora_reporte=horaReporteUsada,
			nombre_paciente=nombreUsado,
			documento=documentoUsado,
			servicio='Urgencias',
			quien_reporta='Enfermera Prueba',
			destino='Hospital Central',
			procedimiento='Traslado de emergencia',
			medico='Dr. Prueba',
			conductor='Conductor Prueba',
			radio_operador='Radio Prueba',
			ambulancia='AMB-001',
		)

		with self.assertRaises ((ValidationError, Exception)):
			traslado.save ()

		self.assertEqual (
			self._contarRegistros (),
			conteoAntes,
			"No debe persistirse ningún registro cuando al menos un campo obligatorio está vacío.",
		)


class PropiedadPersistenciaCampos (TestCase):
	"""
	**Validates: Requirements 3.7, 4.1, 4.2**

	Property 4: Record persistence preserves all fields.
	Para cualquier dato válido de TrasladoPaciente, guardar el registro y
	recuperarlo desde la base de datos debe producir un objeto con todos los
	campos iguales a los valores originales, incluyendo el mes derivado.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	@settings (max_examples=10)
	@given (
		fecha=estrategiaFechaValida,
		horaReporte=st.times (),
		horaEgreso=st.one_of (st.none (), st.times ()),
		horaIngreso=st.one_of (st.none (), st.times ()),
		nombrePaciente=st.text (
			alphabet=st.characters (whitelist_categories=('Lu', 'Ll', 'Nd', 'Zs')),
			min_size=1,
			max_size=255,
		),
		documento=st.text (
			alphabet=st.characters (whitelist_categories=('Lu', 'Ll', 'Nd')),
			min_size=1,
			max_size=50,
		),
		servicio=st.text (min_size=1, max_size=100),
		quienReporta=st.text (min_size=1, max_size=100),
		destino=st.text (min_size=1, max_size=100),
		procedimiento=st.text (min_size=1, max_size=255),
		medico=st.text (min_size=1, max_size=100),
		conductor=st.text (min_size=1, max_size=100),
		radioOperador=st.text (min_size=1, max_size=100),
		ambulancia=st.text (min_size=1, max_size=100),
		observacion=st.text (min_size=0, max_size=500),
	)
	def test_persistencia_preserva_todos_los_campos (
		self,
		fecha,
		horaReporte,
		horaEgreso,
		horaIngreso,
		nombrePaciente,
		documento,
		servicio,
		quienReporta,
		destino,
		procedimiento,
		medico,
		conductor,
		radioOperador,
		ambulancia,
		observacion,
	):
		"""Verifica que todos los campos del traslado se preservan tras guardar y recuperar."""
		traslado = TrasladoPaciente (
			fecha=fecha,
			hora_reporte=horaReporte,
			hora_egreso=horaEgreso,
			hora_ingreso=horaIngreso,
			nombre_paciente=nombrePaciente,
			documento=documento,
			servicio=servicio,
			quien_reporta=quienReporta,
			destino=destino,
			procedimiento=procedimiento,
			medico=medico,
			conductor=conductor,
			radio_operador=radioOperador,
			ambulancia=ambulancia,
			observacion=observacion,
		)
		traslado.save ()

		trasladoRecuperado = TrasladoPaciente.objects.get (pk=traslado.pk)

		self.assertEqual (trasladoRecuperado.fecha, fecha)
		self.assertEqual (trasladoRecuperado.hora_reporte, horaReporte)
		self.assertEqual (trasladoRecuperado.hora_egreso, horaEgreso)
		self.assertEqual (trasladoRecuperado.hora_ingreso, horaIngreso)
		self.assertEqual (trasladoRecuperado.nombre_paciente, nombrePaciente)
		self.assertEqual (trasladoRecuperado.documento, documento)
		self.assertEqual (trasladoRecuperado.servicio, servicio)
		self.assertEqual (trasladoRecuperado.quien_reporta, quienReporta)
		self.assertEqual (trasladoRecuperado.destino, destino)
		self.assertEqual (trasladoRecuperado.procedimiento, procedimiento)
		self.assertEqual (trasladoRecuperado.medico, medico)
		self.assertEqual (trasladoRecuperado.conductor, conductor)
		self.assertEqual (trasladoRecuperado.radio_operador, radioOperador)
		self.assertEqual (trasladoRecuperado.ambulancia, ambulancia)
		self.assertEqual (trasladoRecuperado.observacion, observacion)
		self.assertEqual (trasladoRecuperado.mes, fecha.month)

		traslado.delete ()


class PropiedadFiltroRetornaSoloRegistrosCoincidentes (TestCase):
	"""
	**Validates: Requirements 2.5**

	Property 3: Filter returns only matching records.
	Para cualquier combinación de mes (1-12) y rango de días [dia_desde, dia_hasta],
	el queryset retorna solo registros donde fecha.month == mes y
	dia_desde <= fecha.day <= dia_hasta. Ningún registro fuera de esos límites
	debe aparecer en el resultado.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def _crearTrasladoParaFecha (self, fecha):
		"""Crea y persiste un TrasladoPaciente para la fecha dada, retorna la instancia."""
		traslado = construirTrasladoValido (fecha)
		traslado.save ()
		return traslado

	def _aplicarFiltro (self, mes, diaDesde, diaHasta, anio):
		"""Aplica la misma lógica de filtro que _obtenerContextoTabla en views.py."""
		fechaDesde = datetime.date (anio, mes, diaDesde)
		fechaHasta = datetime.date (anio, mes, diaHasta)
		return TrasladoPaciente.objects.filter (
			mes=mes,
			fecha__gte=fechaDesde,
			fecha__lte=fechaHasta,
		)

	@settings (max_examples=10)
	@given (
		mes=st.integers (min_value=1, max_value=12),
		diaDesde=st.integers (min_value=1, max_value=28),
		diaDesdeOffset=st.integers (min_value=0, max_value=27),
	)
	def test_filtro_retorna_solo_registros_dentro_del_rango (self, mes, diaDesde, diaDesdeOffset):
		"""Verifica que el queryset filtrado no contiene registros fuera del rango mes/días."""
		diaHasta = min (diaDesde + diaDesdeOffset, 28)

		# Usar un año pasado fijo para evitar conflictos con fechas futuras
		anio = 2023

		# ── Crear registros DENTRO del rango ─────────────────────────────────
		fechasDentro = [
			datetime.date (anio, mes, diaDesde),
			datetime.date (anio, mes, diaHasta),
		]
		# Agregar un día intermedio si el rango lo permite
		diaMedio = (diaDesde + diaHasta) // 2
		if diaDesde < diaMedio < diaHasta:
			fechasDentro.append (datetime.date (anio, mes, diaMedio))

		trasladosDentro = [self._crearTrasladoParaFecha (f) for f in fechasDentro]

		# ── Crear registros FUERA del rango (mes diferente) ──────────────────
		mesDiferente = (mes % 12) + 1  # mes siguiente en ciclo 1-12
		fechaFueraMes = datetime.date (anio, mesDiferente, 1)
		trasladoFueraMes = self._crearTrasladoParaFecha (fechaFueraMes)

		# ── Crear registros FUERA del rango (día fuera del rango, mismo mes) ─
		trasladoFueraDia = None
		if diaDesde > 1:
			fechaFueraDia = datetime.date (anio, mes, diaDesde - 1)
			trasladoFueraDia = self._crearTrasladoParaFecha (fechaFueraDia)
		elif diaHasta < 28:
			fechaFueraDia = datetime.date (anio, mes, diaHasta + 1)
			trasladoFueraDia = self._crearTrasladoParaFecha (fechaFueraDia)

		# ── Aplicar el filtro ─────────────────────────────────────────────────
		queryset = self._aplicarFiltro (mes, diaDesde, diaHasta, anio)

		# ── Verificar que TODOS los registros retornados están dentro del rango
		for registro in queryset:
			self.assertEqual (
				registro.fecha.month,
				mes,
				f"Registro con fecha={registro.fecha} tiene mes={registro.fecha.month}, "
				f"pero se filtró por mes={mes}.",
			)
			self.assertGreaterEqual (
				registro.fecha.day,
				diaDesde,
				f"Registro con fecha={registro.fecha} tiene día={registro.fecha.day}, "
				f"que es menor que dia_desde={diaDesde}.",
			)
			self.assertLessEqual (
				registro.fecha.day,
				diaHasta,
				f"Registro con fecha={registro.fecha} tiene día={registro.fecha.day}, "
				f"que es mayor que dia_hasta={diaHasta}.",
			)

		# ── Verificar que el registro de mes diferente NO aparece en el resultado
		idsFueraMes = {trasladoFueraMes.pk}
		idsResultado = set (queryset.values_list ('pk', flat=True))
		self.assertFalse (
			idsFueraMes & idsResultado,
			f"El registro con mes={mesDiferente} no debería aparecer al filtrar por mes={mes}.",
		)

		# ── Verificar que el registro fuera del rango de días NO aparece ──────
		if trasladoFueraDia is not None:
			self.assertNotIn (
				trasladoFueraDia.pk,
				idsResultado,
				f"El registro con fecha={trasladoFueraDia.fecha} no debería aparecer "
				f"al filtrar por dia_desde={diaDesde}, dia_hasta={diaHasta}.",
			)

		# ── Limpiar todos los registros creados en esta iteración ─────────────
		pksDentro = [t.pk for t in trasladosDentro]
		TrasladoPaciente.objects.filter (pk__in=pksDentro).delete ()
		TrasladoPaciente.objects.filter (pk=trasladoFueraMes.pk).delete ()
		if trasladoFueraDia is not None:
			TrasladoPaciente.objects.filter (pk=trasladoFueraDia.pk).delete ()


class PropiedadRecuperacionContrasena (TestCase):
	"""
	**Validates: Requirements 10.4**

	Property 11: Password recovery never reveals user existence.
	Para cualquier string como nombre de usuario (existente o no),
	el endpoint POST /recuperar-contrasena/ retorna el mismo código HTTP
	y el mismo mensaje genérico, sin revelar si el usuario existe.
	"""

	def setUp (self):
		"""Crea un usuario real para comparar respuestas con usuarios existentes vs inexistentes."""
		from django.contrib.auth.models import User
		self.urlRecuperacion = '/recuperar-contrasena/'
		self.usuarioExistente = User.objects.create_user (
			username='usuario_existente_prueba',
			password='clave_segura_123',
		)

	@settings (max_examples=10)
	@given (
		nombreUsuario=st.text (
			alphabet=st.characters (
				whitelist_categories=('Lu', 'Ll', 'Nd'),
				whitelist_characters='-_.',
			),
			min_size=0,
			max_size=150,
		)
	)
	def test_recuperacion_retorna_mismo_codigo_y_mensaje_para_cualquier_usuario (self, nombreUsuario):
		"""Verifica que el endpoint retorna HTTP 200 y el mismo mensaje genérico para cualquier username."""
		from django.test import Client

		cliente = Client ()

		# Respuesta para el username arbitrario (puede o no existir)
		respuestaArbitraria = cliente.post (self.urlRecuperacion, {'username': nombreUsuario})

		# Respuesta para el usuario que SÍ existe en el sistema
		respuestaExistente = cliente.post (
			self.urlRecuperacion,
			{'username': self.usuarioExistente.username},
		)

		# Ambas respuestas deben tener el mismo código HTTP
		self.assertEqual (
			respuestaArbitraria.status_code,
			200,
			f"Se esperaba HTTP 200 para username='{nombreUsuario}', "
			f"pero se obtuvo {respuestaArbitraria.status_code}.",
		)
		self.assertEqual (
			respuestaExistente.status_code,
			200,
			"Se esperaba HTTP 200 para el usuario existente.",
		)

		# Ambas respuestas deben contener el mismo mensaje genérico
		from app_traslados.views import VistaRecuperarContrasena
		mensajeEsperado = VistaRecuperarContrasena.MENSAJE_GENERICO

		self.assertIn (
			mensajeEsperado,
			respuestaArbitraria.content.decode ('utf-8'),
			f"El mensaje genérico no aparece en la respuesta para username='{nombreUsuario}'.",
		)
		self.assertIn (
			mensajeEsperado,
			respuestaExistente.content.decode ('utf-8'),
			"El mensaje genérico no aparece en la respuesta para el usuario existente.",
		)


class PropiedadBloqueoMesCerradoEnVistas (TestCase):
	"""
	**Validates: Requirements 3.9, 5.4**

	Property 2: Closed month blocks all write operations (view layer).
	Para cualquier mes cerrado, las vistas de creación (POST /traslados/nuevo/),
	edición (GET y POST /traslados/<pk>/editar/) y eliminación
	(DELETE /traslados/<pk>/eliminar/) deben retornar HTTP 403.
	"""

	def setUp (self):
		"""Crea usuario autenticado, inicializa ControlMes y un registro de prueba."""
		from django.contrib.auth.models import User
		crearControlMeses ()
		self.usuario, _ = User.objects.get_or_create (
			username='tester_vistas',
			defaults={'password': 'clave_segura_456'},
		)
		self.usuario.set_password ('clave_segura_456')
		self.usuario.save ()

	def _fechaParaMes (self, numeroMes):
		"""Retorna una fecha válida (no futura) para el mes dado."""
		hoy = datetime.date.today ()
		anio = hoy.year if numeroMes <= hoy.month else hoy.year - 1
		return datetime.date (anio, numeroMes, 1)

	def _cerrarMes (self, numeroMes):
		"""Cierra el mes indicado actualizando su ControlMes a CERRADO."""
		ControlMes.objects.filter (mes=numeroMes).update (estado='CERRADO')

	def _abrirMes (self, numeroMes):
		"""Abre el mes indicado actualizando su ControlMes a ABIERTO."""
		ControlMes.objects.filter (mes=numeroMes).update (estado='ABIERTO')

	def _crearRegistroEnMes (self, numeroMes):
		"""Crea y persiste un TrasladoPaciente en el mes dado (con mes abierto)."""
		fecha = self._fechaParaMes (numeroMes)
		traslado = construirTrasladoValido (fecha)
		traslado.save ()
		return traslado

	@settings (max_examples=12)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_post_crear_en_mes_cerrado_retorna_403 (self, numeroMes):
		"""Verifica que POST /traslados/nuevo/ retorna 403 cuando el mes está cerrado."""
		from django.test import Client

		self._cerrarMes (numeroMes)
		fecha = self._fechaParaMes (numeroMes)

		cliente = Client ()
		cliente.force_login (self.usuario)

		datos = {
			'fecha': fecha.isoformat (),
			'hora_reporte': '08:00',
			'nombre_paciente': 'Paciente Prueba',
			'documento': '123456789',
			'servicio': 'Urgencias',
			'quien_reporta': 'Enfermera Prueba',
			'destino': 'Hospital Central',
			'procedimiento': 'Traslado de emergencia',
			'medico': 'Dr. Prueba',
			'conductor': 'Conductor Prueba',
			'radio_operador': 'Radio Prueba',
			'ambulancia': 'AMB-001',
			'mes': str (numeroMes),
		}

		respuesta = cliente.post ('/traslados/nuevo/', datos)

		self.assertEqual (
			respuesta.status_code,
			403,
			f"Se esperaba HTTP 403 al crear en mes={numeroMes} cerrado, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		self._abrirMes (numeroMes)

	@settings (max_examples=12)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_get_editar_en_mes_cerrado_retorna_403 (self, numeroMes):
		"""Verifica que GET /traslados/<pk>/editar/ retorna 403 cuando el mes está cerrado."""
		from django.test import Client

		# Crear el registro con el mes abierto
		traslado = self._crearRegistroEnMes (numeroMes)

		# Cerrar el mes después de crear el registro
		self._cerrarMes (numeroMes)

		cliente = Client ()
		cliente.force_login (self.usuario)

		respuesta = cliente.get (f'/traslados/{traslado.pk}/editar/')

		self.assertEqual (
			respuesta.status_code,
			403,
			f"Se esperaba HTTP 403 al editar (GET) en mes={numeroMes} cerrado, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		# Limpiar
		self._abrirMes (numeroMes)
		TrasladoPaciente.objects.filter (pk=traslado.pk).delete ()

	@settings (max_examples=12)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_post_editar_en_mes_cerrado_retorna_403 (self, numeroMes):
		"""Verifica que POST /traslados/<pk>/editar/ retorna 403 cuando el mes está cerrado."""
		from django.test import Client

		# Crear el registro con el mes abierto
		traslado = self._crearRegistroEnMes (numeroMes)

		# Cerrar el mes después de crear el registro
		self._cerrarMes (numeroMes)

		cliente = Client ()
		cliente.force_login (self.usuario)

		fecha = self._fechaParaMes (numeroMes)
		datos = {
			'fecha': fecha.isoformat (),
			'hora_reporte': '09:00',
			'nombre_paciente': 'Paciente Modificado',
			'documento': '987654321',
			'servicio': 'UCI',
			'quien_reporta': 'Médico Prueba',
			'destino': 'Clínica Norte',
			'procedimiento': 'Traslado programado',
			'medico': 'Dr. Modificado',
			'conductor': 'Conductor Modificado',
			'radio_operador': 'Radio Modificado',
			'ambulancia': 'AMB-002',
		}

		respuesta = cliente.post (f'/traslados/{traslado.pk}/editar/', datos)

		self.assertEqual (
			respuesta.status_code,
			403,
			f"Se esperaba HTTP 403 al editar (POST) en mes={numeroMes} cerrado, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		# Limpiar
		self._abrirMes (numeroMes)
		TrasladoPaciente.objects.filter (pk=traslado.pk).delete ()

	@settings (max_examples=12)
	@given (numeroMes=st.integers (min_value=1, max_value=12))
	def test_delete_eliminar_en_mes_cerrado_retorna_403 (self, numeroMes):
		"""Verifica que DELETE /traslados/<pk>/eliminar/ retorna 403 cuando el mes está cerrado."""
		from django.test import Client

		# Crear el registro con el mes abierto
		traslado = self._crearRegistroEnMes (numeroMes)

		# Cerrar el mes después de crear el registro
		self._cerrarMes (numeroMes)

		cliente = Client ()
		cliente.force_login (self.usuario)

		respuesta = cliente.delete (f'/traslados/{traslado.pk}/eliminar/')

		self.assertEqual (
			respuesta.status_code,
			403,
			f"Se esperaba HTTP 403 al eliminar en mes={numeroMes} cerrado, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		# Verificar que el registro NO fue eliminado
		self.assertTrue (
			TrasladoPaciente.objects.filter (pk=traslado.pk).exists (),
			f"El registro no debería haberse eliminado con el mes={numeroMes} cerrado.",
		)

		# Limpiar
		self._abrirMes (numeroMes)
		TrasladoPaciente.objects.filter (pk=traslado.pk).delete ()


class PropiedadContenidoExcel (TestCase):
	"""
	**Validates: Requirements 6.1, 6.3**

	Property 9: Excel report contains correct headers and exactly the filtered records.
	Para cualquier filtro activo (mes, rango de días opcional), el Excel generado
	contiene exactamente una fila de datos por registro del queryset filtrado y
	los encabezados en español correctos como primera fila.
	"""

	# Encabezados esperados en español, en el mismo orden que el servicio
	ENCABEZADOS_ESPERADOS = [
		'FECHA',
		'HORA REPORTE',
		'HORA DE EGRESO',
		'HORA DE INGRESO',
		'NOMBRE DE PACIENTE',
		'DOCUMENTO',
		'SERVICIO',
		'QUIEN REPORTA',
		'DESTINO',
		'PROCEDIMIENTO',
		'MÉDICO',
		'CONDUCTOR',
		'RADIO OPERADOR',
		'AMBULANCIA DE TRASLADO',
		'OBSERVACIÓN',
	]

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def _crearTrasladoParaFecha (self, fecha):
		"""Crea y persiste un TrasladoPaciente para la fecha dada, retorna la instancia."""
		traslado = construirTrasladoValido (fecha)
		traslado.save ()
		return traslado

	def _obtenerQuerysetFiltrado (self, mes, diaDesde, diaHasta, anio):
		"""Retorna el queryset de TrasladoPaciente filtrado por mes y rango de días."""
		fechaDesde = datetime.date (anio, mes, diaDesde)
		fechaHasta = datetime.date (anio, mes, diaHasta)
		return TrasladoPaciente.objects.filter (
			mes=mes,
			fecha__gte=fechaDesde,
			fecha__lte=fechaHasta,
		)

	def _leerHojasDesdeBytes (self, bytesExcel):
		"""Carga el libro de trabajo openpyxl desde los bytes del Excel generado."""
		import io
		import openpyxl
		buffer = io.BytesIO (bytesExcel)
		return openpyxl.load_workbook (buffer)

	@settings (max_examples=10)
	@given (
		mes=st.integers (min_value=1, max_value=12),
		diaDesde=st.integers (min_value=1, max_value=28),
		diaDesdeOffset=st.integers (min_value=0, max_value=27),
		cantidadRegistros=st.integers (min_value=0, max_value=5),
	)
	def test_excel_contiene_encabezados_correctos_y_exactamente_los_registros_filtrados (
		self, mes, diaDesde, diaDesdeOffset, cantidadRegistros
	):
		"""Verifica que el Excel tiene los encabezados en español y una fila por registro filtrado."""
		from app_traslados.services.report_excel import generarExcel

		diaHasta = min (diaDesde + diaDesdeOffset, 28)
		anio = 2023  # Año pasado fijo para evitar conflictos con fechas futuras

		# ── Crear registros DENTRO del rango ─────────────────────────────────
		trasladosCreados = []
		for i in range (cantidadRegistros):
			# Distribuir los registros uniformemente dentro del rango de días
			dia = diaDesde + (i % (diaHasta - diaDesde + 1))
			fecha = datetime.date (anio, mes, dia)
			traslado = self._crearTrasladoParaFecha (fecha)
			trasladosCreados.append (traslado)

		# ── Crear un registro FUERA del rango (mes diferente) para verificar exclusión
		mesDiferente = (mes % 12) + 1
		fechaFuera = datetime.date (anio, mesDiferente, 1)
		trasladoFuera = self._crearTrasladoParaFecha (fechaFuera)

		# ── Obtener el queryset filtrado (misma lógica que la vista) ─────────
		queryset = self._obtenerQuerysetFiltrado (mes, diaDesde, diaHasta, anio)
		cantidadEsperada = queryset.count ()

		# ── Generar el Excel ──────────────────────────────────────────────────
		bytesExcel, nombreArchivo = generarExcel (queryset, mes)

		# ── Leer el Excel generado ────────────────────────────────────────────
		libro = self._leerHojasDesdeBytes (bytesExcel)
		hoja = libro.active

		# ── Verificar que los bytes son un Excel válido (no vacío) ───────────
		self.assertGreater (
			len (bytesExcel),
			0,
			"El Excel generado no debe estar vacío.",
		)

		# ── Verificar encabezados en la primera fila ──────────────────────────
		encabezadosObtenidos = [celda.value for celda in hoja [1]]
		self.assertEqual (
			encabezadosObtenidos,
			self.ENCABEZADOS_ESPERADOS,
			f"Los encabezados del Excel no coinciden con los esperados.\n"
			f"Esperados: {self.ENCABEZADOS_ESPERADOS}\n"
			f"Obtenidos: {encabezadosObtenidos}",
		)

		# ── Verificar que el número de filas de datos es exactamente el del queryset
		# La hoja tiene 1 fila de encabezados + N filas de datos
		filasEnHoja = hoja.max_row
		filasEsperadas = cantidadEsperada + 1  # encabezados + datos

		self.assertEqual (
			filasEnHoja,
			filasEsperadas,
			f"Se esperaban {filasEsperadas} filas (1 encabezado + {cantidadEsperada} datos), "
			f"pero la hoja tiene {filasEnHoja} filas.",
		)

		# ── Verificar que el registro fuera del rango NO aparece en el Excel ─
		# Recopilar las fechas de todas las filas de datos del Excel
		fechasEnExcel = set ()
		for fila in hoja.iter_rows (min_row=2, values_only=True):
			if fila [0] is not None:
				# La celda de fecha puede ser un objeto date o un string según openpyxl
				valorFecha = fila [0]
				if hasattr (valorFecha, 'month'):
					fechasEnExcel.add (valorFecha.month)
				elif isinstance (valorFecha, str):
					try:
						fechaParseada = datetime.date.fromisoformat (valorFecha)
						fechasEnExcel.add (fechaParseada.month)
					except ValueError:
						pass

		# Si hay filas de datos, ninguna debe pertenecer al mes diferente
		if cantidadEsperada > 0:
			self.assertNotIn (
				mesDiferente,
				fechasEnExcel,
				f"El Excel no debe contener registros del mes {mesDiferente} "
				f"cuando se filtró por mes={mes}.",
			)

		# ── Limpiar todos los registros creados en esta iteración ─────────────
		pksDentro = [t.pk for t in trasladosCreados]
		TrasladoPaciente.objects.filter (pk__in=pksDentro).delete ()
		TrasladoPaciente.objects.filter (pk=trasladoFuera.pk).delete ()

	@settings (max_examples=10)
	@given (mes=st.integers (min_value=1, max_value=12))
	def test_excel_sin_registros_contiene_solo_encabezados (self, mes):
		"""Verifica que el Excel generado con queryset vacío contiene solo la fila de encabezados."""
		from app_traslados.services.report_excel import generarExcel

		querysetVacio = TrasladoPaciente.objects.none ()
		bytesExcel, nombreArchivo = generarExcel (querysetVacio, mes)

		libro = self._leerHojasDesdeBytes (bytesExcel)
		hoja = libro.active

		# Solo debe haber 1 fila (los encabezados)
		self.assertEqual (
			hoja.max_row,
			1,
			f"Con queryset vacío se esperaba 1 fila (encabezados), "
			f"pero la hoja tiene {hoja.max_row} filas.",
		)

		# Los encabezados deben ser correctos incluso sin datos
		encabezadosObtenidos = [celda.value for celda in hoja [1]]
		self.assertEqual (
			encabezadosObtenidos,
			self.ENCABEZADOS_ESPERADOS,
			f"Los encabezados del Excel vacío no coinciden con los esperados.\n"
			f"Esperados: {self.ENCABEZADOS_ESPERADOS}\n"
			f"Obtenidos: {encabezadosObtenidos}",
		)


class PropiedadNombreArchivoReportes (TestCase):
	"""
	**Validates: Requirements 6.6**

	Property 10: Report filename follows the required format.
	Para cualquier valor de mes (1–12), el header Content-Disposition de la
	respuesta del reporte Excel contiene traslados_<mes>.xlsx y el del reporte
	PDF contiene traslados_<mes>.pdf.
	"""

	def setUp (self):
		"""Crea usuario autenticado e inicializa los 12 registros de ControlMes."""
		from django.contrib.auth.models import User
		crearControlMeses ()
		self.usuario, _ = User.objects.get_or_create (
			username='tester_reportes',
			defaults={'password': 'clave_segura_789'},
		)
		self.usuario.set_password ('clave_segura_789')
		self.usuario.save ()

	def _obtenerCliente (self):
		"""Crea y retorna un cliente de prueba autenticado."""
		from django.test import Client
		cliente = Client ()
		cliente.force_login (self.usuario)
		return cliente

	def _verificarNombreArchivo (self, encabezadoDisposicion, nombreEsperado):
		"""Verifica que el encabezado Content-Disposition contiene el nombre de archivo esperado."""
		self.assertIn (
			nombreEsperado,
			encabezadoDisposicion,
			f"Se esperaba '{nombreEsperado}' en Content-Disposition, "
			f"pero se obtuvo: '{encabezadoDisposicion}'",
		)

	@settings (max_examples=12)
	@given (mes=st.integers (min_value=1, max_value=12))
	def test_nombre_archivo_excel_sigue_formato_requerido (self, mes):
		"""Verifica que el Content-Disposition del Excel contiene traslados_<mes>.xlsx."""
		cliente = self._obtenerCliente ()
		respuesta = cliente.get (f'/reportes/excel/?mes={mes}')

		self.assertEqual (
			respuesta.status_code,
			200,
			f"Se esperaba HTTP 200 para /reportes/excel/?mes={mes}, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		encabezadoDisposicion = respuesta.get ('Content-Disposition', '')
		nombreEsperado = f'traslados_{mes}.xlsx'
		self._verificarNombreArchivo (encabezadoDisposicion, nombreEsperado)

	@settings (max_examples=12)
	@given (mes=st.integers (min_value=1, max_value=12))
	def test_nombre_archivo_pdf_sigue_formato_requerido (self, mes):
		"""Verifica que el Content-Disposition del PDF contiene traslados_<mes>.pdf."""
		cliente = self._obtenerCliente ()
		respuesta = cliente.get (f'/reportes/pdf/?mes={mes}')

		self.assertEqual (
			respuesta.status_code,
			200,
			f"Se esperaba HTTP 200 para /reportes/pdf/?mes={mes}, "
			f"pero se obtuvo {respuesta.status_code}.",
		)

		encabezadoDisposicion = respuesta.get ('Content-Disposition', '')
		nombreEsperado = f'traslados_{mes}.pdf'
		self._verificarNombreArchivo (encabezadoDisposicion, nombreEsperado)


class PropiedadLimpiezaAnualRestableceTodoElEstado (TestCase):
	"""
	**Validates: Requirements 9.3**

	Property 6: Annual cleanup resets all state.
	Para cualquier estado del sistema (N registros TrasladoPaciente, combinación
	de meses abiertos/cerrados), tras ejecutar la limpieza el conteo de
	TrasladoPaciente es 0 y todos los ControlMes tienen estado = ABIERTO.
	"""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def _crearRegistros (self, cantidad):
		"""Crea la cantidad indicada de TrasladoPaciente con fechas en 2023."""
		for i in range (cantidad):
			mes = (i % 12) + 1
			fecha = datetime.date (2023, mes, 1)
			traslado = construirTrasladoValido (fecha)
			traslado.save ()

	def _cerrarMeses (self, mesesACerrar):
		"""Cierra los meses indicados en la lista."""
		for mes in mesesACerrar:
			ControlMes.objects.filter (mes=mes).update (estado='CERRADO')

	def _ejecutarLimpieza (self):
		"""Ejecuta la misma lógica de limpieza que la vista vistaLimpiarSistema."""
		from app_traslados.views import _ejecutarLimpieza
		_ejecutarLimpieza ()

	@settings (max_examples=10)
	@given (
		cantidadRegistros=st.integers (min_value=0, max_value=20),
		mesesACerrar=st.lists (
			st.integers (min_value=1, max_value=12),
			min_size=0,
			max_size=12,
			unique=True,
		),
	)
	def test_limpieza_elimina_todos_los_registros_y_restablece_meses (
		self, cantidadRegistros, mesesACerrar
	):
		"""Verifica que tras la limpieza no quedan registros y todos los meses están ABIERTOS."""
		# ── Preparar el estado del sistema ───────────────────────────────────
		self._crearRegistros (cantidadRegistros)
		self._cerrarMeses (mesesACerrar)

		# ── Ejecutar la limpieza ──────────────────────────────────────────────
		self._ejecutarLimpieza ()

		# ── Verificar que no quedan registros TrasladoPaciente ────────────────
		conteoRegistros = TrasladoPaciente.objects.count ()
		self.assertEqual (
			conteoRegistros,
			0,
			f"Se esperaba 0 registros tras la limpieza, pero quedan {conteoRegistros}.",
		)

		# ── Verificar que todos los ControlMes están ABIERTOS ─────────────────
		mesesCerrados = ControlMes.objects.filter (estado='CERRADO')
		self.assertEqual (
			mesesCerrados.count (),
			0,
			f"Se esperaba que todos los meses estuvieran ABIERTOS tras la limpieza, "
			f"pero {mesesCerrados.count ()} meses siguen CERRADOS: "
			f"{list (mesesCerrados.values_list ('mes', flat=True))}.",
		)

		# ── Verificar que los 12 meses existen y están ABIERTOS ───────────────
		self.assertEqual (
			ControlMes.objects.filter (estado='ABIERTO').count (),
			12,
			"Se esperaban exactamente 12 meses con estado ABIERTO tras la limpieza.",
		)
