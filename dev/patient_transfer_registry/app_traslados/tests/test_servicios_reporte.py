"""
Tests unitarios para los servicios de generación de reportes.

Verifica que generarExcel() y generarPdf() producen archivos válidos
con los encabezados correctos, el nombre de archivo esperado y el
comportamiento correcto cuando el queryset está vacío.
"""

import datetime
import io

import openpyxl
from django.test import TestCase

from app_traslados.models import ControlMes, TrasladoPaciente
from app_traslados.services.report_excel import ENCABEZADOS, generarExcel
from app_traslados.services.report_pdf import generarPdf


def crearControlMeses ():
	"""Crea las 12 filas de ControlMes con estado ABIERTO si no existen."""
	for numeroMes in range (1, 13):
		ControlMes.objects.get_or_create (
			mes=numeroMes,
			defaults={'estado': 'ABIERTO'},
		)


def crearTrasladoValido (fecha=None):
	"""Crea y persiste un TrasladoPaciente con todos los campos requeridos."""
	if fecha is None:
		fecha = datetime.date.today ()
	traslado = TrasladoPaciente (
		fecha=fecha,
		hora_reporte=datetime.time (8, 0),
		hora_egreso=datetime.time (9, 0),
		hora_ingreso=datetime.time (10, 0),
		nombre_paciente='Paciente Prueba',
		documento='123456789',
		servicio='Urgencias',
		quien_reporta='Enfermera Prueba',
		destino='Hospital Central',
		procedimiento='Traslado de emergencia',
		medico='Dr. Prueba',
		conductor='Conductor Prueba',
		radio_operador='Radio Prueba',
		ambulancia='AMB-001',
		observacion='Sin novedad',
	)
	traslado.save ()
	return traslado


class PruebasServicioExcel (TestCase):
	"""Tests unitarios para el servicio generarExcel()."""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def test_generarExcelRetornaBytesValidos (self):
		"""Verifica que generarExcel() retorna bytes no vacíos de un .xlsx válido."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarExcel (queryset, 1)

		self.assertIsInstance (bytesArchivo, bytes)
		self.assertGreater (len (bytesArchivo), 0)

	def test_generarExcelRetornaXlsxLegible (self):
		"""Verifica que los bytes retornados pueden ser leídos por openpyxl."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarExcel (queryset, 1)

		libro = openpyxl.load_workbook (io.BytesIO (bytesArchivo))
		self.assertIsNotNone (libro)

	def test_generarExcelContieneEncabezadosEnEspanol (self):
		"""Verifica que la primera fila del Excel contiene los encabezados en español correctos."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarExcel (queryset, 1)

		libro = openpyxl.load_workbook (io.BytesIO (bytesArchivo))
		hoja = libro.active
		encabezadosObtenidos = [celda.value for celda in hoja [1]]

		self.assertEqual (encabezadosObtenidos, ENCABEZADOS)

	def test_generarExcelSinRegistrosSoloTieneEncabezados (self):
		"""Verifica que con queryset vacío la hoja tiene solo la fila de encabezados."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarExcel (queryset, 5)

		libro = openpyxl.load_workbook (io.BytesIO (bytesArchivo))
		hoja = libro.active

		# Solo debe existir la fila 1 (encabezados); la fila 2 no debe tener datos
		self.assertEqual (hoja.max_row, 1)

	def test_generarExcelConRegistrosAgregaFilasDeDatos (self):
		"""Verifica que cada registro del queryset genera exactamente una fila de datos."""
		fecha = datetime.date.today ()
		crearTrasladoValido (fecha)
		crearTrasladoValido (fecha)

		queryset = TrasladoPaciente.objects.filter (mes=fecha.month)
		bytesArchivo, _ = generarExcel (queryset, fecha.month)

		libro = openpyxl.load_workbook (io.BytesIO (bytesArchivo))
		hoja = libro.active

		# Fila 1 = encabezados; filas 2 y 3 = datos
		self.assertEqual (hoja.max_row, 3)

	def test_generarExcelNombreArchivoSigueFormato (self):
		"""Verifica que el nombre de archivo retornado sigue el formato traslados_<mes>.xlsx."""
		queryset = TrasladoPaciente.objects.none ()
		for mes in [1, 6, 12]:
			_, nombreArchivo = generarExcel (queryset, mes)
			self.assertEqual (nombreArchivo, f'traslados_{mes}.xlsx')

	def test_generarExcelEncabezadosTienenTodosLosCampos (self):
		"""Verifica que los encabezados cubren todos los campos de TrasladoPaciente."""
		self.assertEqual (len (ENCABEZADOS), 15)
		camposEsperados = [
			'FECHA',
			'HORA REPORTE',
			'HORA DE EGRESO',
			'HORA DE INGRESO',
			'NOMBRE DE PACIENTE',
			'DOCUMENTO',
			'SERVICIO',
			'QUIEN REPORTA',
			'DESTINO',
			'PROCEDIMIENTO',
			'MÉDICO',
			'CONDUCTOR',
			'RADIO OPERADOR',
			'AMBULANCIA DE TRASLADO',
			'OBSERVACIÓN',
		]
		self.assertEqual (ENCABEZADOS, camposEsperados)


class PruebasServicioPdf (TestCase):
	"""Tests unitarios para el servicio generarPdf()."""

	def setUp (self):
		"""Inicializa los 12 registros de ControlMes antes de cada prueba."""
		crearControlMeses ()

	def test_generarPdfRetornaBytesNoVacios (self):
		"""Verifica que generarPdf() retorna bytes no vacíos."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarPdf (queryset, 1)

		self.assertIsInstance (bytesArchivo, bytes)
		self.assertGreater (len (bytesArchivo), 0)

	def test_generarPdfConRegistrosRetornaBytesNoVacios (self):
		"""Verifica que generarPdf() con registros retorna bytes no vacíos."""
		fecha = datetime.date.today ()
		crearTrasladoValido (fecha)

		queryset = TrasladoPaciente.objects.filter (mes=fecha.month)
		bytesArchivo, _ = generarPdf (queryset, fecha.month)

		self.assertIsInstance (bytesArchivo, bytes)
		self.assertGreater (len (bytesArchivo), 0)

	def test_generarPdfNombreArchivoSigueFormato (self):
		"""Verifica que el nombre de archivo retornado sigue el formato traslados_<mes>.pdf."""
		queryset = TrasladoPaciente.objects.none ()
		for mes in [1, 6, 12]:
			_, nombreArchivo = generarPdf (queryset, mes)
			self.assertEqual (nombreArchivo, f'traslados_{mes}.pdf')

	def test_generarPdfRetornaSignaturaPdf (self):
		"""Verifica que los bytes retornados comienzan con la firma de un archivo PDF."""
		queryset = TrasladoPaciente.objects.none ()
		bytesArchivo, _ = generarPdf (queryset, 3)

		# Los archivos PDF comienzan con el encabezado %PDF
		self.assertTrue (bytesArchivo.startswith (b'%PDF'))
